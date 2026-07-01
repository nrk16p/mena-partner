"""
Migrate contract financial details from ผ่อนเงินดาวน์ sheet.

Column mapping (0-indexed):
  0: contractCode
  4: totalPrice (ลูกหนี้คงค้าง = original financed amount)
  6: downPayment (ยอดเงินดาวน์)
  8: netDebt (รวมยอดหนี้สุทธิ = totalPrice - downPayment)
  9: totalInstallments (จำนวนงวด)
  10: monthlyInstallment (เดือนละ @)
  11: startDate (เริ่มผ่อนงวดแรก)

Also migrates repair installment info from หนี้ผ่อนชำระใหม่ sheet:
  0: contractCode
  3: repairInstallmentDown (initial repair payment)
  4: repairDebtBalance
  5: repairTotal
  6: repairMonthlyInstallment
  7: repairInstallmentCount

Run: python3 scripts/seed-contract-financials.py  (from project root)
"""
import openpyxl
from pymongo import MongoClient, UpdateOne
from datetime import datetime, timezone

MONGO_URI  = "mongodb+srv://adminbew:K879w5XpBm3QL046@mn-mongodb-ops-2c8032e6.mongo.ondigitalocean.com/terminus?authSource=admin"
DB_NAME    = "mena_partner"
EXCEL_PATH = "Rev.o1 Payroll รถร่วม Mixer มิ.ย. 69.xlsx"

client = MongoClient(MONGO_URI)
db     = client[DB_NAME]
now    = datetime.now(timezone.utc)

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

def to_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default

def to_int(val, default=0):
    try:
        return int(val) if val is not None else default
    except (TypeError, ValueError):
        return default

# ─── 1. ผ่อนเงินดาวน์ → contracts financial fields ────────────────────────────
ws = wb["ผ่อนเงินดาวน์"]
rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
data_rows = [r for r in rows if r[0] and str(r[0]).startswith("MT")]

ops = []
for row in data_rows:
    code              = str(row[0]).strip()
    total_price       = to_float(row[4])
    down_payment      = to_float(row[6])
    total_installments = to_int(row[9])
    monthly_installment = to_float(row[10])
    start_date        = row[11]  # datetime or None

    update = {
        "totalPrice":          total_price,
        "downPayment":         down_payment,
        "totalInstallments":   total_installments,
        "monthlyInstallment":  monthly_installment,
        "updatedAt":           now,
    }
    if isinstance(start_date, datetime):
        update["startDate"]    = start_date
        update["contractDate"] = start_date

    ops.append(UpdateOne({"contractCode": code}, {"$set": update}))

if ops:
    r = db["contracts"].bulk_write(ops)
    print(f"ผ่อนเงินดาวน์ → contracts: matched={r.matched_count}  modified={r.modified_count}")
else:
    print("ผ่อนเงินดาวน์: no rows found")

# ─── 2. หนี้ผ่อนชำระใหม่ → contracts repair installment fields ────────────────
ws2 = wb["หนี้ผ่อนชำระใหม่"]
rows2 = [r for r in ws2.iter_rows(values_only=True) if any(c is not None for c in r)]
data2 = [r for r in rows2 if r[0] and str(r[0]).startswith("MT")]

ops2 = []
for row in data2:
    code                     = str(row[0]).strip()
    repair_down              = to_float(row[3])   # initial repair installment down
    repair_balance           = to_float(row[4])   # remaining debt
    repair_total             = to_float(row[5])   # total repair debt
    repair_monthly           = to_float(row[6])   # monthly installment
    repair_count             = to_int(row[7])     # number of installments

    ops2.append(UpdateOne({"contractCode": code}, {"$set": {
        "repairInstallmentDown":    repair_down,
        "repairDebtBalance":        repair_balance,
        "repairInstallmentTotal":   repair_total,
        "repairMonthlyInstallment": repair_monthly,
        "repairInstallmentCount":   repair_count,
        "updatedAt":                now,
    }}))

if ops2:
    r2 = db["contracts"].bulk_write(ops2)
    print(f"หนี้ผ่อนชำระใหม่ → contracts: matched={r2.matched_count}  modified={r2.modified_count}")
else:
    print("หนี้ผ่อนชำระใหม่: no rows found")

# ─── Spot-check ────────────────────────────────────────────────────────────────
sample = db["contracts"].find_one({"contractCode": "MTL003"}, {
    "_id": 0, "contractCode": 1, "totalPrice": 1, "downPayment": 1,
    "monthlyInstallment": 1, "totalInstallments": 1, "startDate": 1,
    "repairMonthlyInstallment": 1,
})
print(f"\nSpot-check MTL003: {sample}")

client.close()
print("\n✓ Contract financials migration complete.")
