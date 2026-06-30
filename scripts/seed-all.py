"""
Seed script: imports contracts, drivers, and promo_config from Excel.

Sources:
  - NetPay sheet   → contracts + drivers (102 records)
  - โปรโมชั่น ซ่อม+PM sheet → promo_config (237 records, matched by licensePlate)

Run: python3 scripts/seed-all.py
"""
import openpyxl
from pymongo import MongoClient, UpdateOne
from datetime import datetime

MONGO_URI  = "mongodb+srv://adminbew:K879w5XpBm3QL046@mn-mongodb-ops-2c8032e6.mongo.ondigitalocean.com/terminus?authSource=admin"
DB_NAME    = "mena_partner"
EXCEL_PATH = "Rev.o1 Payroll รถร่วม Mixer มิ.ย. 69.xlsx"

client = MongoClient(MONGO_URI)
db     = client[DB_NAME]
now    = datetime.utcnow()

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

# ─── 1. Contracts + Drivers from NetPay ───────────────────────────────────────
ws_net = wb["NetPay"]
rows   = [r for r in ws_net.iter_rows(values_only=True) if any(c is not None for c in r)]
data   = rows[3:]   # skip 3 header rows

contract_ops = []
driver_ops   = []
plate_to_code = {}   # built here for promo mapping

for row in data:
    seq, code, buyer, driver, plate, truck = row[0], row[1], row[2], row[3], row[4], row[5]
    if not code or not isinstance(code, str):
        continue
    code  = str(code).strip()
    plate = str(plate).strip() if plate else ""
    truck = str(truck).strip() if truck else ""

    if plate:
        plate_to_code[plate] = code

    contract_doc = {
        "contractCode":  code,
        "buyerName":     str(buyer).strip() if buyer else "",
        "driverName":    str(driver).strip() if driver else "",
        "licensePlate":  plate,
        "truckNumber":   truck,
        "status":        "active",
        # fields not yet available from Excel — left blank
        "phone":               "",
        "plant":               "",
        "vehicleBrand":        "",
        "accountNumber":       "",
        "totalPrice":          0,
        "downPayment":         0,
        "monthlyInstallment":  0,
        "totalInstallments":   0,
        "notes":               "",
    }
    contract_ops.append(UpdateOne(
        {"contractCode": code},
        {"$set": contract_doc, "$setOnInsert": {"createdAt": now, "contractDate": now, "startDate": now, "updatedAt": now}},
        upsert=True
    ))

    driver_doc = {
        "contractCode":  code,
        "buyerName":     contract_doc["buyerName"],
        "driverName":    contract_doc["driverName"],
        "truckNumber":   truck,
        "licensePlate":  plate,
        "phone":         "",
        "plant":         "",
        "status":        "active",
    }
    driver_ops.append(UpdateOne(
        {"contractCode": code},
        {"$set": driver_doc, "$setOnInsert": {"createdAt": now}},
        upsert=True
    ))

r1 = db["contracts"].bulk_write(contract_ops)
print(f"contracts — upserted: {r1.upserted_count}, modified: {r1.modified_count}")

r2 = db["drivers"].bulk_write(driver_ops)
print(f"drivers   — upserted: {r2.upserted_count}, modified: {r2.modified_count}")

# ─── 2. promo_config from โปรโมชั่น ซ่อม+PM ──────────────────────────────────
ws_promo  = wb["โปรโมชั่น ซ่อม+PM"]
promo_rows = [r for r in ws_promo.iter_rows(values_only=True) if any(c is not None for c in r)]
promo_data = promo_rows[1:]   # skip header

promo_ops  = []
not_found  = []

for row in promo_data:
    plate, repair_budget, pm_oil_cost, annual_pm_cap = row[0], row[1], row[2], row[3]
    if not plate:
        continue
    plate = str(plate).strip()
    code  = plate_to_code.get(plate)

    doc = {
        "licensePlate": plate,
        "repairBudget": float(repair_budget or 0),
        "pmOilCost":    float(pm_oil_cost or 0),
        "annualPmCap":  float(annual_pm_cap or 0),
    }
    if code:
        doc["contractCode"] = code
    else:
        not_found.append(plate)

    promo_ops.append(UpdateOne(
        {"licensePlate": plate},
        {"$set": doc, "$setOnInsert": {"createdAt": now}},
        upsert=True
    ))

r3 = db["promo_config"].bulk_write(promo_ops)
print(f"promo_config — upserted: {r3.upserted_count}, modified: {r3.modified_count}")

if not_found:
    print(f"\n⚠  {len(not_found)} plates in promo sheet have no matching contract (stored without contractCode):")
    for p in not_found[:10]:
        print(f"   {p}")
    if len(not_found) > 10:
        print(f"   ...and {len(not_found)-10} more")

client.close()
print("\n✓ Seed complete.")
