"""
Seed monthly_adjustments collection from 'รับอื่นๆ' and 'หักอื่นๆ' sheets.
Stores summary totals of other income and other deductions per contract per month.

รับอื่นๆ column mapping (0-indexed):
  0:  contractCode
  16: otherIncomeWHT   (รวมรับอื่นๆ หัก WHT)
  21: otherIncomeNoWHT (รวมรับอื่นๆ ไม่หัก WHT)

หักอื่นๆ column mapping (0-indexed):
  0:  contractCode
  10: otherDeductionWHT   (รวมหักอื่นๆ WHT)
  16: otherDeductionNoWHT (รวมหักอื่นๆ ไม WHT)

Run: python3 scripts/seed-monthly-adjustments.py  (from project root)
"""
import openpyxl
from pymongo import MongoClient, UpdateOne
from datetime import datetime, timezone

MONGO_URI  = "mongodb+srv://adminbew:K879w5XpBm3QL046@mn-mongodb-ops-2c8032e6.mongo.ondigitalocean.com/terminus?authSource=admin"
DB_NAME    = "mena_partner"
EXCEL_PATH = "Rev.o1 Payroll รถร่วม Mixer มิ.ย. 69.xlsx"
MONTH      = "2026-05"

client = MongoClient(MONGO_URI)
db     = client[DB_NAME]
now    = datetime.now(timezone.utc)

def to_float(v):
    try:
        return float(str(v).replace(",", "")) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# --- รับอื่นๆ ---
ws_in = wb["รับอื่นๆ"]
rows_in = [r for r in ws_in.iter_rows(values_only=True) if any(c is not None for c in r)]
income_data = {
    str(r[0]).strip(): r
    for r in rows_in
    if r[0] and str(r[0]).strip().startswith("MT")
}
print(f"รับอื่นๆ rows: {len(income_data)}")

# --- หักอื่นๆ ---
ws_out = wb["หักอื่นๆ"]
rows_out = [r for r in ws_out.iter_rows(values_only=True) if any(c is not None for c in r)]
deduct_data = {
    str(r[0]).strip(): r
    for r in rows_out
    if r[0] and str(r[0]).strip().startswith("MT")
}
print(f"หักอื่นๆ rows: {len(deduct_data)}")

all_codes = set(income_data.keys()) | set(deduct_data.keys())
print(f"Unique contracts: {len(all_codes)}")

ops = []
for code in sorted(all_codes):
    in_row  = income_data.get(code, [])
    out_row = deduct_data.get(code, [])

    other_income_wht     = to_float(in_row[16]) if len(in_row) > 16 else 0.0
    other_income_no_wht  = to_float(in_row[21]) if len(in_row) > 21 else 0.0
    other_deduct_wht     = to_float(out_row[10]) if len(out_row) > 10 else 0.0
    other_deduct_no_wht  = to_float(out_row[16]) if len(out_row) > 16 else 0.0

    doc = {
        "contractCode":      code,
        "month":             MONTH,
        "otherIncomeWHT":    round(other_income_wht, 2),
        "otherIncomeNoWHT":  round(other_income_no_wht, 2),
        "otherDeductWHT":    round(other_deduct_wht, 2),
        "otherDeductNoWHT":  round(other_deduct_no_wht, 2),
        "updatedAt":         now,
    }
    ops.append(UpdateOne(
        {"contractCode": code, "month": MONTH},
        {"$set": doc, "$setOnInsert": {"createdAt": now}},
        upsert=True
    ))

if ops:
    r = db["monthly_adjustments"].bulk_write(ops)
    print(f"Upserted: {r.upserted_count}  Modified: {r.modified_count}")

# Spot-check MTL003 — expect otherIncomeWHT=4150.86, otherIncomeNoWHT=0, otherDeductWHT=0
sample = db["monthly_adjustments"].find_one({"contractCode": "MTL003", "month": MONTH})
if sample:
    print(f"\nSpot-check MTL003: incWHT={sample['otherIncomeWHT']} incNoWHT={sample['otherIncomeNoWHT']} dedWHT={sample['otherDeductWHT']} dedNoWHT={sample['otherDeductNoWHT']}")

client.close()
print("Done.")
